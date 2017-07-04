# -*- coding: utf-8 -*-

import csv
import argparse
import datetime
import openpyxl

NEWLINE = ''
SEMICOLON = ';'
FILE_PATH = '/home/aries/PycharmProjects/BuildingStatistics/Распределение студентов 1 курса на Stepik.xlsx'
INPUT_FILE = 'sourse.csv'
INITIAL_DATA = 'Начальные данные, {}'
REPORT = 'Сводка, {}'
REPORT_DATE = 'Сводка, {}'
USER_ID = 'user_id'
LAST_NAME = 'last_name'
FIRST_NAME = 'first_name'
TOTAL = 'total'
MAX_SCORE = 'Максимальный балл'
PERCENTAGE_OF_TOTAL = '% от максимального балла'
SCORED_0_POINTS = 'Студенты, которые набрали 0 баллов'
GAINED_LESS_40_PERCENT = 'Список студентов, которые набрали < 40%'
GAINED_MORE_40_LESS_70_PERCENT = 'Список студентов, которые набрали > 39% и < 71%'
GAINED_MORE_70_PERCENT = 'Список студентов, которые набрали > 70%'
SCORED_MAX_SCORE = 'Студенты, которые набрали максимальный балл ({})'

def pars():
    parser = argparse.ArgumentParser()
    parser.add_argument('-m', '--max', type=int, nargs=1, help='Max Score')
    parser.add_argument('-n', '--name', nargs=1, help='Module Name')
    max_score = parser.parse_args()
    module_name = parser.parse_args()
    return max_score.max[0], module_name.name[0]

def filling_table():
    user_id = []
    last_name = []
    first_name = []
    total = []
    input_file = open(INPUT_FILE, 'rb')
    rdr = csv.reader(input_file)
    next(rdr)
    for user in rdr :
        id = NEWLINE
        last = NEWLINE
        first = NEWLINE
        tot = NEWLINE
        flag_id = True
        flag_last = False
        flag_first = False
        flag_total = False

        for i in user[0]:
            if flag_id:
                if i == SEMICOLON:
                    flag_id = False
                    flag_last = True
                    continue
                else:
                    id += i
            if flag_last:
                if i == SEMICOLON:
                    flag_last = False
                    flag_first = True
                    continue
                else:
                    last += i
            if flag_first:
                if i == SEMICOLON:
                    flag_first = False
                    flag_total = True
                    continue
                else:
                    first += i
            if flag_total:
                if i == SEMICOLON:
                    flag_id = False
                else:
                    tot += i
        user_id.append(id)
        last_name.append(last)
        first_name.append(first)
        total.append(tot)

    input_file.close()
    return user_id, last_name, first_name, total

def summary_table(sheet, user_id, last_name, first_name, total, value):
    for i in range(0, len(user_id)):
        sheet['A{}'.format(i + 2)] = user_id[i]
        sheet['B{}'.format(i + 2)] = last_name[i]
        sheet['C{}'.format(i + 2)] = first_name[i]
        sheet['D{}'.format(i + 2)] = int(total[i])
        sheet['E{}'.format(i + 2)] = int(int(total[i]) * 100 / value)

def number_points(sheet, user_id, last_name, first_name, total, value, last_line_index):
    curr_last_line_index = last_line_index
    count = 1
    for i in range(0, len(user_id)):
        if int(total[i]) == value:
            sheet['A{}'.format(count + last_line_index)] = user_id[i]
            sheet['B{}'.format(count + last_line_index)] = last_name[i]
            sheet['C{}'.format(count + last_line_index)] = first_name[i]
            curr_last_line_index = count + last_line_index
            count += 1
    return (curr_last_line_index + 2)

def percentage(sheet, user_id, last_name, first_name, total, max_value, last_line_index, a, b):
    curr_last_line_index = last_line_index
    count = 1
    for i in range(0, len(total)):
        if a <= int(total[i]) * 100 / max_value <= b:
            sheet['A{}'.format(count + last_line_index)] = user_id[i]
            sheet['B{}'.format(count + last_line_index)] = last_name[i]
            sheet['C{}'.format(count + last_line_index)] = first_name[i]
            sheet['D{}'.format(count + last_line_index)] = int(total[i])
            sheet['E{}'.format(count + last_line_index)] = int(int(total[i]) * 100 / max_value)
            curr_last_line_index = count + last_line_index
            count += 1

    return (curr_last_line_index + 2)

def first_table(wb, module_name, max_score, user_id, last_name, first_name, total):
    sheet = wb.create_sheet(INITIAL_DATA.format(str(module_name)).decode('utf8'), 0)
    sheet['A1'] = USER_ID
    sheet['B1'] = LAST_NAME
    sheet['C1'] = FIRST_NAME
    sheet['D1'] = TOTAL
    sheet['E1'] = PERCENTAGE_OF_TOTAL
    sheet['G1'] = MAX_SCORE
    sheet['H1'] = max_score

    summary_table(sheet, user_id, last_name, first_name, total, max_score)

def second_table(wb, name, user_id, last_name, first_name, total, max_score):
    sheet = wb.create_sheet(REPORT_DATE.format(str(name) + ' ' + str(datetime.datetime.now().strftime('%d %B'))).decode('utf8'), 1)
    sheet['A1'] = GAINED_LESS_40_PERCENT
    num = percentage(sheet, user_id, last_name, first_name, total, max_score, 1, 0, 39)
    sheet['A{}'.format(num)] = GAINED_MORE_40_LESS_70_PERCENT
    num = percentage(sheet, user_id, last_name, first_name, total, max_score, num, 40, 70)
    sheet['A{}'.format(num)] = GAINED_MORE_70_PERCENT
    percentage(sheet, user_id, last_name, first_name, total, max_score, num, 70, 100)

def third_table(wb, name, user_id, last_name, first_name, total, max_score):
    sheet = wb.create_sheet(REPORT.format(str(name)).decode('utf8'), 2)
    sheet['A1'] = SCORED_0_POINTS
    num = number_points(sheet, user_id, last_name, first_name, total, 0, 1)
    sheet['A{}'.format(num)] = SCORED_MAX_SCORE.format(max_score)
    number_points(sheet, user_id, last_name, first_name, total, max_score, num)

def main():
    if __name__ == '__main__':
        max_score, module_name = pars()
        wb = openpyxl.Workbook()
        user_id, last_name, first_name, total = filling_table()
        first_table(wb, module_name, max_score, user_id, last_name, first_name, total)
        second_table(wb, module_name, user_id, last_name, first_name, total, max_score)
        third_table(wb, module_name, user_id, last_name, first_name, total, max_score)
        wb.save(FILE_PATH)

main()
